
#coding=gbk

import openpyxl


from openpyxl import load_workbook

# MineModule  LoginModule  PostModule MessageCenterModule MomentModule MineSetModule ConfigModule MessageListModule GoFishingModule SecretStoryModule
# ['��������', '��½ע��', '��̬', '˽��', '˲��Ⱥ', '��������', 'ͨ��', '��Ϣ�б�', '����', '����']

wb = load_workbook('./��������Ӣ�Ķ���.xlsx')
localModuleFilePath = "files/ConfigModule.strings"

#��ȡ������--Sheet
# �������sheet������
print(str(wb.sheetnames))
# ����sheet���ֻ��sheet
a_sheet = wb['ͨ��']
# ���sheet��
print(a_sheet.title)
# ��õ�ǰ������ʾ��sheet, Ҳ������wb.get_active_sheet()
sheet = wb.active
def writeFileHead():
    with open(localModuleFilePath, 'a') as file_object:
        file_object.write("/*\n " +localModuleFilePath+"\n Pods\n Created by liguanglei on 2023/3/29.\n\n*/\n\n")

def writeRowTofile(key, value):
    with open(localModuleFilePath, 'a') as file_object:
        file_object.write("\"" + key+"\" = \"" + value + "\";\n")


writeFileHead()
for row in a_sheet.values:
    key = str(row[0])
    chineseValue = str(row[1])
    chineseValue = chineseValue.replace("{count}","%s")
    yinhaoStr = "\\\""
    chineseValue = chineseValue.replace("\"", yinhaoStr)
    if key != "None" and key != "" and key != "Keyֵ" and chineseValue != "None":
        print(key + "=" + chineseValue + "��ע��" + str(row[2]))
        writeRowTofile(key, chineseValue)



